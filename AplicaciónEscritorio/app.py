from datetime import datetime
import json
import os
import pandas as pd
import serial 
from kivy.app import App
from kivy.uix.screenmanager import Screen
from kivy.lang import Builder
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from openpyxl import load_workbook
from kivy.metrics import dp
from kivymd.app import MDApp
from kivymd.uix.datatables import MDDataTable
import matplotlib.pyplot as plt
from kivy.uix.image import Image
from cryptography.fernet import Fernet

def cargar_clave():
    return open('clave.key', 'rb').read()

def encriptar(nom_archivo, clave):
    f = Fernet(clave)
    with open(nom_archivo, 'rb') as file:
        archivo_info = file.read()
    encrypted_data = f.encrypt(archivo_info)
    with open(nom_archivo, 'wb') as file:
        file.write(encrypted_data)

def desencriptar(nom_archivo, clave):
    f = Fernet(clave)
    with open(nom_archivo, 'rb') as file:
        encrypted_data = file.read()
    decrypted_data = f.decrypt(encrypted_data)
    with open(nom_archivo, 'wb') as file:
        file.write(decrypted_data)

clave = cargar_clave()
desencriptar("usuarios.json", clave)
desencriptar("registro_pacientes.xlsx", clave)

usuarios_registrados = "usuarios.json"

def guardar_usuarios(usuarios):
    with open(usuarios_registrados, "w") as f:
        json.dump(usuarios, f)

def cargar_usuarios():
    if os.path.exists(usuarios_registrados):
        with open(usuarios_registrados, "r") as f:
            contenido = f.read()
            if contenido.strip():
                return json.loads(contenido)
    return {}

usuarios = cargar_usuarios()

def mostrar_advertencia(mensaje):
    popup = Popup(
        title='ERROR',
        content=Label(text=mensaje),
        size_hint=(None, None),
        size=(400, 200))
    popup.open()

class PrincipalScreen(Screen):
    pass

class VisualizacionScreen(Screen):
    def actualizar_nombre(self, nombre_completo):
        self.ids.nombre_paciente_visualizacion.text = f' {nombre_completo.replace("_", " ") }'
        self.tabla(nombre_completo)
    def tabla(self, nombre_completo):
        nombre_apellidos = nombre_completo.split()
        if len(nombre_apellidos) > 3:
            nombre_con_guion = nombre_apellidos[1] + "_" + nombre_apellidos[2] + "_" + nombre_apellidos[3]
        else:
            nombre_con_guion = nombre_apellidos[1] + "_" + nombre_apellidos[2]
        nombre_hoja = nombre_con_guion
        DATA = pd.read_excel("registro_pacientes.xlsx", sheet_name=nombre_hoja)
        cols = DATA.columns.values
        values = DATA.values
        self.data_tables = MDDataTable(
            use_pagination=True,
            column_data=[(col, dp(30)) for col in cols],
            row_data=values)
        contenedor = self.ids.get("contenedor_tabla")
        contenedor.clear_widgets()
        contenedor.add_widget(self.data_tables)

class VisualizargraficaScreen(Screen):
    def actualizar_nombre(self, nombre_completo):
        self.ids.nombre_paciente_visualizacion.text = f' {nombre_completo.replace("_", " ") }'
        self.grafica(nombre_completo)
    def grafica(self, nombre_completo):
        nombre_apellidos = nombre_completo.split()
        if len(nombre_apellidos) > 3:
            nombre_con_guion = nombre_apellidos[1] + "_" + nombre_apellidos[2] + "_" + nombre_apellidos[3]
        else:
            nombre_con_guion = nombre_apellidos[1] + "_" + nombre_apellidos[2]
        nombre_hoja = nombre_con_guion
        df = pd.read_excel("registro_pacientes.xlsx", sheet_name=nombre_hoja)
        fecha = df['Fecha']
        dedos = ['Pulgar(Kg)','Indice(Kg)','Corazón(Kg)','Anular(Kg)','Meñique(Kg)']
        plt.figure(figsize=(10, 6))
        for dedo in dedos:
            plt.plot(fecha, df[dedo], marker='o', label=dedo)
        plt.title("Fuerza por dedo (Kg)")
        plt.xlabel("Fecha")
        plt.ylabel("Fuerza (Kg)")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.savefig("grafica_dedos.png")
        plt.close()
        contenedor = self.ids.contenedor_grafica
        contenedor.clear_widgets()
        imagen = Image()
        imagen.source = "grafica_dedos.png"
        imagen.reload()
        contenedor.add_widget(imagen)

class PantallaPacienteScreen(Screen):
    def ir_a_visualizacion(self, nombre_completo):
        app = App.get_running_app()
        visualizacion_screen = app.root.get_screen("PantallaVisualizacion")
        visualizacion_screen.actualizar_nombre(nombre_completo)
        app.root.current = "PantallaVisualizacion"
    def ir_a_visualizar(self, nombre_completo):
        app = App.get_running_app()
        visualizacion_screen = app.root.get_screen("visualizargrafica")
        visualizacion_screen.actualizar_nombre(nombre_completo)
        app.root.current = "visualizargrafica"
    def actualizar_nombre(self, nombre_completo):
        self.ids.paciente_nombre_input.text = f'Paciente: {nombre_completo.replace("_", " ")}'
    def medir(self, nombre_completo):
        archivo_excel = "registro_pacientes.xlsx"
        nombre_apellidos = nombre_completo.split()
        if len(nombre_apellidos) > 3:
            nombre_con_guion = nombre_apellidos[1] + "_" + nombre_apellidos[2] + "_" + nombre_apellidos[3]
        else:
            nombre_con_guion = nombre_apellidos[1] + "_" + nombre_apellidos[2]
        nombre_hoja = nombre_con_guion
        datos = []
        if not os.path.exists(archivo_excel):
            mostrar_advertencia("El archivo Excel no existe.")
        arduino = serial.Serial('COM5', 9600, timeout=1)
        if not arduino:
            mostrar_advertencia('Placa arduino no detectada')
        medicion = False
        while not medicion:
            linea = arduino.readline().decode('utf-8').strip()
            if "Pulgar:" in linea:
                partes = linea.split("|")
                valores = {}
                for parte in partes:
                    if ":" in parte:
                        nombre, valor = parte.split(":")
                        nombre = nombre.strip()
                        valor = valor.replace("Kg", "").strip()
                        valores[nombre] = valor
                fechaHora = datetime.now().strftime("%Y-%m-%d %H:%M")
                datos.append({
                    "Fecha": fechaHora,
                    "Pulgar (Kg)": valores.get("Pulgar", ""),
                    "Índice (Kg)": valores.get("Indice", ""),
                    "Corazón (Kg)": valores.get("Corazon", ""),
                    "Anular (Kg)": valores.get("Anular", ""),
                    "Meñique (Kg)": valores.get("Menique", "")})
                medicion = True                 
        arduino.close()
        if datos:
            df = pd.DataFrame(datos)
            libro = load_workbook("registro_pacientes.xlsx")
            hoja = libro[nombre_hoja]
            fila_inicio = hoja.max_row
            with pd.ExcelWriter("registro_pacientes.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=nombre_hoja, index=False, header=False, startrow=fila_inicio)
            app = App.get_running_app()
            visualizacion_screen = app.root.get_screen("PantallaVisualizacion")
            visualizacion_screen.actualizar_nombre(nombre_completo)
        else:
            mostrar_advertencia("No se recogieron datos válidos.")

class RegistroPacienteScreen(Screen):
    def registrar_datos_paciente(self):
        nombre = self.ids.paciente_nombre_input.text
        apellidos = self.ids.paciente_apellidos_input.text
        if not nombre or not apellidos:
            mostrar_advertencia("Es obligatorio rellenar todos los datos.")
        app = App.get_running_app()
        app.root.current = "nueva"
        archivo_excel = "registro_pacientes.xlsx"
        nombre_hoja = f"{nombre}_{apellidos}".replace(" ", "_")
        if not os.path.exists(archivo_excel):
            with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
                df = pd.DataFrame(columns=["Fecha","Pulgar(Kg)","Indice(Kg)","Corazón(Kg)","Anular(Kg)","Meñique(Kg)"]) 
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        else:
            with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='a') as writer:
                book = load_workbook(archivo_excel)
                if nombre_hoja not in book.sheetnames:
                    df = pd.DataFrame(columns=["Fecha","Pulgar(Kg)","Indice(Kg)","Corazón(Kg)","Anular(Kg)","Meñique(Kg)"])
                    df.to_excel(writer, sheet_name=nombre_hoja, index=False)

class NuevaScreen(Screen):
    def on_enter(self): 
        archivo = load_workbook("registro_pacientes.xlsx")
        pacientes = list(archivo.sheetnames)
        self.ids.paciente_spinner.values = pacientes
    def seleccionar_paciente(self, nombre_completo):
        if nombre_completo != "Seleccionar paciente":
            pantalla_paciente = self.manager.get_screen("PantallaPaciente")
            pantalla_paciente.actualizar_nombre(nombre_completo)
            self.manager.current = "PantallaPaciente"
class IniciarScreen(Screen):
    def verificar_login(self):
        usuario = self.ids.login_usuario.text
        contraseña = self.ids.login_contrasena.text
        if usuario in usuarios and usuarios[usuario]["contraseña"] == contraseña:
            app = App.get_running_app()
            app.nombre_usuario = usuario
            nueva_screen = app.root.get_screen("nueva")
            nueva_screen.ids.label_nombre.text = f" Profesional: {usuarios[usuario]['nombre']} {usuarios[usuario]['apellidos']} "
            app.root.current = "nueva"
        else:
            mostrar_advertencia("Usuario o contraseña incorrectos.")

class RegistroScreen(Screen):
    def registrar_usuario(self):
        nombre = self.ids.nombre_input.text
        apellidos = self.ids.apellidos_input.text
        usuario = self.ids.usuario_input.text
        contraseña = self.ids.contrasena_input.text
        if not usuario or not contraseña:
            mostrar_advertencia("Usuario y contraseña son obligatorios.")
        elif usuario in usuarios:
            mostrar_advertencia("El usuario ya está registrado.")
        else:
            usuarios[usuario] = {"nombre": nombre,"apellidos": apellidos,"contraseña": contraseña}
            guardar_usuarios(usuarios)
            App.get_running_app().root.current = "principal"

class MiApp(MDApp):
    def build(self):
        return Builder.load_file("interfaz.kv")
    def on_stop(self):
        clave = cargar_clave()
        encriptar("usuarios.json", clave)
        encriptar("registro_pacientes.xlsx", clave)
MiApp().run()
